import io
from datetime import datetime
from importlib import import_module
from importlib.util import find_spec
from typing import TYPE_CHECKING, List, Optional, Tuple

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from markupsafe import Markup

if TYPE_CHECKING:  # pragma: no cover - assists static analysis only
    from openpyxl import Workbook as WorkbookType
    from openpyxl import load_workbook as LoadWorkbookType
else:  # pragma: no cover - runtime alias without importing optional dependency
    WorkbookType = object
    LoadWorkbookType = object

_OPENPYXL_CACHE: Tuple[Optional[WorkbookType], Optional[LoadWorkbookType], bool]
_OPENPYXL_CACHE = (None, None, False)


def _ensure_openpyxl() -> Tuple[Optional[WorkbookType], Optional[LoadWorkbookType], bool]:
    """Lazily import openpyxl, returning its key entry points when available."""

    global _OPENPYXL_CACHE
    workbook_cls, load_workbook_fn, has_attempted = _OPENPYXL_CACHE
    if has_attempted:
        return workbook_cls, load_workbook_fn, workbook_cls is not None and load_workbook_fn is not None

    if find_spec("openpyxl") is None:  # pragma: no cover - optional dependency may be absent
        _OPENPYXL_CACHE = (None, None, True)
        return None, None, False

    module = import_module("openpyxl")

    workbook_cls = getattr(module, "Workbook", None)
    load_workbook_fn = getattr(module, "load_workbook", None)
    _OPENPYXL_CACHE = (workbook_cls, load_workbook_fn, True)
    return workbook_cls, load_workbook_fn, workbook_cls is not None and load_workbook_fn is not None

from ..extensions import db
from ..models import Class, Grade, Reader
from ..utils.pagination import get_page_args


bp = Blueprint("readers", __name__, url_prefix="/readers")


@bp.route("/")
@login_required
def list_readers():
    keyword = request.args.get("q", "").strip()
    grade_id_raw = request.args.get("grade_id", "").strip()
    class_id_raw = request.args.get("class_id", "").strip()
    show_deleted = request.args.get("show_deleted") == "1" if current_user.level == "admin" else False
    grade_id = int(grade_id_raw) if grade_id_raw.isdigit() else None
    class_id = int(class_id_raw) if class_id_raw.isdigit() else None
    page, per_page = get_page_args()
    query = Reader.query
    if show_deleted:
        query = query.filter_by(is_deleted=True)
    else:
        query = query.filter_by(is_deleted=False)
    if keyword:
        like = f"%{keyword}%"
        query = query.filter((Reader.name.like(like)) | (Reader.card_no.like(like)))
    if grade_id is not None or class_id is not None:
        query = query.join(Class, Reader.reader_class)
        if grade_id is not None:
            query = query.join(Grade, Class.grade).filter(Grade.id == grade_id)
        if class_id is not None:
            query = query.filter(Class.id == class_id)
    pagination = query.order_by(Reader.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    grades = (
        Grade.query.filter_by(is_deleted=False)
        .order_by(Grade.name)
        .all()
    )
    classes_query = (
        Class.query.filter_by(is_deleted=False)
        .join(Grade)
        .filter(Grade.is_deleted.is_(False))
    )
    if grade_id is not None:
        classes_query = classes_query.filter(Class.grade_id == grade_id)
    classes = classes_query.order_by(Grade.name, Class.name).all()
    return render_template(
        "readers/list.html",
        readers=pagination.items,
        filters={
            "keyword": keyword,
            "grade_id": grade_id,
            "class_id": class_id,
            "show_deleted": show_deleted,
        },
        pagination=pagination,
        grades=grades,
        classes=classes,
    )


@bp.route("/create", methods=["GET", "POST"])
@login_required
def create_reader():
    classes = (
        Class.query.filter_by(is_deleted=False)
        .join(Grade)
        .filter(Grade.is_deleted.is_(False))
        .order_by(Grade.name, Class.name)
        .all()
    )

    if request.method == "GET":
        return render_template("readers/create.html", classes=classes)

    form = request.form
    card_no = form.get("card_no", "").strip()
    if not card_no:
        flash("读者卡号不能为空", "danger")
        return redirect(url_for("readers.create_reader"))
    existing_active = Reader.query.filter_by(card_no=card_no, is_deleted=False).first()
    if existing_active:
        flash("卡号已存在", "danger")
        return redirect(url_for("readers.create_reader"))

    class_id_value = int(form.get("class_id")) if form.get("class_id") else None
    def _apply_reader_fields(reader: Reader) -> None:
        reader.name = form.get("name", reader.name or "未命名读者")
        reader.sex = form.get("sex")
        reader.phone = form.get("phone")
        reader.cert_type = form.get("cert_type")
        reader.cert_no = form.get("cert_no")
        reader.class_id = class_id_value

    existing_deleted = Reader.query.filter_by(card_no=card_no, is_deleted=True).first()
    if existing_deleted:
        _apply_reader_fields(existing_deleted)
        existing_deleted.is_deleted = False
        existing_deleted.status = "active"
        db.session.commit()
        flash("读者创建成功", "success")
        return redirect(url_for("readers.list_readers"))

    reader = Reader(
        card_no=card_no,
        name=form.get("name", "未命名读者"),
        sex=form.get("sex"),
        phone=form.get("phone"),
        cert_type=form.get("cert_type"),
        cert_no=form.get("cert_no"),
        class_id=class_id_value,
    )
    db.session.add(reader)
    db.session.commit()
    flash("读者创建成功", "success")
    return redirect(url_for("readers.list_readers"))


@bp.route("/<int:reader_id>/update", methods=["POST"])
@login_required
def update_reader(reader_id: int):
    reader = Reader.query.get_or_404(reader_id)
    form = request.form
    reader.card_no = form.get("card_no", reader.card_no)
    reader.name = form.get("name", reader.name)
    reader.sex = form.get("sex")
    reader.phone = form.get("phone")
    reader.cert_type = form.get("cert_type")
    reader.cert_no = form.get("cert_no")
    reader.class_id = int(form.get("class_id")) if form.get("class_id") else None
    db.session.commit()
    flash("读者信息已更新", "success")
    return redirect(url_for("readers.list_readers"))


@bp.route("/<int:reader_id>/delete", methods=["POST"])
@login_required
def delete_reader(reader_id: int):
    if current_user.level != "admin":
        flash("只有管理员可以执行删除操作", "danger")
        return redirect(url_for("readers.list_readers"))
    reader = Reader.query.get_or_404(reader_id)
    reader.is_deleted = True
    db.session.commit()
    flash("读者已删除", "success")
    return redirect(url_for("readers.list_readers"))


@bp.route("/bulk-delete", methods=["POST"])
@login_required
def bulk_delete_readers():
    if current_user.level != "admin":
        flash("只有管理员可以执行删除操作", "danger")
        return redirect(url_for("readers.list_readers"))
    reader_ids = request.form.getlist("reader_ids")
    selected_ids = []
    for reader_id in reader_ids:
        try:
            selected_ids.append(int(reader_id))
        except (TypeError, ValueError):
            continue

    next_url = request.form.get("next")

    if not selected_ids:
        flash("请选择要删除的读者", "warning")
        return redirect(next_url or url_for("readers.list_readers"))

    readers = (
        Reader.query.filter(Reader.id.in_(selected_ids))
        .filter(Reader.is_deleted.is_(False))
        .all()
    )
    if not readers:
        flash("未找到可删除的读者", "warning")
        return redirect(next_url or url_for("readers.list_readers"))

    for reader in readers:
        reader.is_deleted = True
    db.session.commit()
    flash(f"已删除 {len(readers)} 名读者", "success")
    return redirect(next_url or url_for("readers.list_readers"))


@bp.route("/import", methods=["POST"])
@login_required
def import_readers():
    _, load_workbook_fn, has_openpyxl = _ensure_openpyxl()
    if not has_openpyxl or load_workbook_fn is None:
        flash("当前环境缺少 openpyxl 依赖，无法导入读者数据。请先运行 pip install openpyxl。", "danger")
        return redirect(url_for("readers.list_readers"))

    if request.form.get("template_confirmed") != "1":
        flash("请先下载导入模板并确认后再上传数据。", "warning")
        return redirect(url_for("readers.list_readers"))

    file = request.files.get("file")
    if not file:
        flash("请选择Excel文件", "danger")
        return redirect(url_for("readers.list_readers"))

    try:
        wb = load_workbook_fn(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created_count = 0
        restored_count = 0
        skipped: List[str] = []

        for index, row in enumerate(rows, start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            card_no = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            if not card_no:
                skipped.append(f"第{index}行: 卡号不能为空")
                continue

            existing_active = Reader.query.filter_by(card_no=card_no, is_deleted=False).first()
            if existing_active:
                skipped.append(f"第{index}行(卡号 {card_no}): 卡号已存在")
                continue

            name_value = str(row[1]).strip() if len(row) > 1 and row[1] else "未命名读者"
            phone_value = str(row[2]).strip() if len(row) > 2 and row[2] else None
            sex_value = str(row[3]).strip() if len(row) > 3 and row[3] else None
            grade_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            class_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            klass = None
            if class_name:
                class_query = Class.query.filter_by(name=class_name, is_deleted=False)
                if grade_name:
                    grade = Grade.query.filter_by(name=grade_name, is_deleted=False).first()
                    if grade:
                        klass = class_query.filter_by(grade_id=grade.id).first()
                if klass is None:
                    klass = class_query.join(Grade).filter(Grade.is_deleted.is_(False)).first()
            elif grade_name:
                grade = Grade.query.filter_by(name=grade_name, is_deleted=False).first()
                if grade:
                    klass = (
                        Class.query.filter_by(is_deleted=False, grade_id=grade.id)
                        .order_by(Class.name)
                        .first()
                    )

            def _assign(reader: Reader) -> None:
                reader.name = name_value or reader.name or "未命名读者"
                reader.phone = phone_value
                reader.sex = sex_value
                reader.class_id = klass.id if klass else None

            existing_deleted = Reader.query.filter_by(card_no=card_no, is_deleted=True).first()
            if existing_deleted:
                _assign(existing_deleted)
                existing_deleted.is_deleted = False
                existing_deleted.status = "active"
                restored_count += 1
                continue

            reader = Reader(card_no=card_no)
            _assign(reader)
            db.session.add(reader)
            created_count += 1

        db.session.commit()

        success_parts = []
        if created_count:
            success_parts.append(f"成功导入 {created_count} 条读者记录")
        if restored_count:
            success_parts.append(f"恢复 {restored_count} 条已删除读者记录")
        if success_parts:
            flash("，".join(success_parts), "success")

        if skipped:
            preview = skipped[:100]
            if len(skipped) > 100:
                preview.append(f"……共 {len(skipped)} 条错误，仅显示前100条")
            flash(Markup("部分数据导入失败：<br>" + "<br>".join(preview)), "danger")
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        flash(f"导入失败: {exc}", "danger")

    return redirect(url_for("readers.list_readers"))


@bp.route("/import-template")
@login_required
def download_reader_template():
    workbook_cls, _, has_openpyxl = _ensure_openpyxl()
    if not has_openpyxl or workbook_cls is None:
        flash("当前环境缺少 openpyxl 依赖，无法生成模板。请先运行 pip install openpyxl。", "danger")
        return redirect(url_for("readers.list_readers"))

    wb = workbook_cls()
    ws = wb.active
    ws.title = "读者信息"
    ws.append(["卡号", "姓名", "电话", "性别", "年级", "班级"])

    grades = Grade.query.filter_by(is_deleted=False).order_by(Grade.name).all()
    classes = (
        Class.query.filter_by(is_deleted=False)
        .join(Grade)
        .filter(Grade.is_deleted.is_(False))
        .order_by(Grade.name, Class.name)
        .all()
    )

    data_sheet = wb.create_sheet("年级班级信息")
    data_sheet.append(["年级", "班级"])
    for idx, grade in enumerate(grades, start=2):
        data_sheet.cell(row=idx, column=1, value=grade.name)
    for idx, klass in enumerate(classes, start=2):
        grade_name = klass.grade.name if klass.grade else ""
        display_name = f"{grade_name}{klass.name}" if grade_name else klass.name
        data_sheet.cell(row=idx, column=2, value=display_name)
    data_sheet.sheet_state = "hidden"

    try:
        dv_module = import_module("openpyxl.worksheet.datavalidation")
        DataValidation = getattr(dv_module, "DataValidation", None)
        utils_module = import_module("openpyxl.utils")
        get_column_letter = getattr(utils_module, "get_column_letter", None)
        quote_sheetname_fn = getattr(utils_module, "quote_sheetname", None)
    except Exception:  # noqa: BLE001 - fallback when openpyxl structure changes
        DataValidation = None
        get_column_letter = None
        quote_sheetname_fn = None

    def _quote_sheet_name(name: str) -> str:
        if quote_sheetname_fn:
            return quote_sheetname_fn(name)
        escaped = name.replace("'", "''")
        return f"'{escaped}'"

    if DataValidation and get_column_letter:
        max_rows = 500
        grade_column = get_column_letter(5)
        class_column = get_column_letter(6)
        if grades:
            grade_list_ref = f"={_quote_sheet_name(data_sheet.title)}!$A$2:$A${len(grades) + 1}"
            grade_range = f"{grade_column}2:{grade_column}{max_rows}"
            grade_validation = DataValidation(type="list", formula1=grade_list_ref, allow_blank=True)
            grade_validation.error = "请选择下拉列表中的年级"
            grade_validation.errorTitle = "无效的年级"
            ws.add_data_validation(grade_validation)
            grade_validation.add(grade_range)
        if classes:
            class_list_ref = f"={_quote_sheet_name(data_sheet.title)}!$B$2:$B${len(classes) + 1}"
            class_range = f"{class_column}2:{class_column}{max_rows}"
            class_validation = DataValidation(type="list", formula1=class_list_ref, allow_blank=True)
            class_validation.error = "请选择下拉列表中的班级"
            class_validation.errorTitle = "无效的班级"
            ws.add_data_validation(class_validation)
            class_validation.add(class_range)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="reader-import-template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/export")
@login_required
def export_readers():
    workbook_cls, _, has_openpyxl = _ensure_openpyxl()
    if not has_openpyxl or workbook_cls is None:
        flash("当前环境缺少 openpyxl 依赖，无法导出读者数据。请先运行 pip install openpyxl。", "danger")
        return redirect(url_for("readers.list_readers"))

    wb = workbook_cls()
    ws = wb.active
    ws.append(["卡号", "姓名", "电话", "性别", "年级", "班级"])
    for reader in Reader.query.filter_by(is_deleted=False).order_by(Reader.name).all():
        ws.append(
            [
                reader.card_no,
                reader.name,
                reader.phone,
                reader.sex,
                reader.reader_class.grade.name if reader.reader_class and reader.reader_class.grade else "",
                reader.reader_class.name if reader.reader_class else "",
            ]
        )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"readers-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/grades")
@login_required
def manage_grades():
    page, per_page = get_page_args()
    pagination = Grade.query.filter_by(is_deleted=False).order_by(Grade.name).paginate(
        page=page, per_page=per_page, error_out=False
    )
    return render_template("readers/grades.html", grades=pagination.items, pagination=pagination)


@bp.route("/grades/create", methods=["GET", "POST"])
@login_required
def create_grade():
    if request.method == "GET":
        return render_template("readers/grades_create.html")

    name = request.form.get("name", "").strip()
    if not name:
        flash("年级名称不能为空", "danger")
        return redirect(url_for("readers.create_grade"))
    grade = Grade(name=name)
    db.session.add(grade)
    db.session.commit()
    flash("年级创建成功", "success")
    return redirect(url_for("readers.manage_grades"))


@bp.route("/grades/<int:grade_id>/delete", methods=["POST"])
@login_required
def delete_grade(grade_id: int):
    if current_user.level != "admin":
        flash("只有管理员可以执行删除操作", "danger")
        return redirect(url_for("readers.manage_grades"))
    grade = Grade.query.get_or_404(grade_id)
    grade.is_deleted = True
    db.session.commit()
    flash("年级已删除", "success")
    return redirect(url_for("readers.manage_grades"))


@bp.route("/classes")
@login_required
def manage_classes():
    page, per_page = get_page_args()
    pagination = (
        Class.query.filter_by(is_deleted=False)
        .join(Grade)
        .filter(Grade.is_deleted.is_(False))
        .order_by(Grade.name, Class.name)
        .paginate(page=page, per_page=per_page, error_out=False)
    )
    grades = Grade.query.filter_by(is_deleted=False).order_by(Grade.name).all()
    return render_template(
        "readers/classes.html",
        classes=pagination.items,
        grades=grades,
        pagination=pagination,
    )


@bp.route("/classes/create", methods=["GET", "POST"])
@login_required
def create_class():
    grades = Grade.query.filter_by(is_deleted=False).order_by(Grade.name).all()
    if request.method == "GET":
        return render_template("readers/classes_create.html", grades=grades)

    grade_id_raw = request.form.get("grade_id")
    names_raw = request.form.get("names", "")
    if not grade_id_raw or not names_raw.strip():
        flash("请选择年级并输入至少一个班级名称", "danger")
        return redirect(url_for("readers.create_class"))

    try:
        grade_id = int(grade_id_raw)
    except ValueError:
        flash("年级选择无效", "danger")
        return redirect(url_for("readers.create_class"))

    grade = Grade.query.filter_by(id=grade_id, is_deleted=False).first()
    if grade is None:
        flash("年级不存在或已被删除", "danger")
        return redirect(url_for("readers.create_class"))

    names = [name.strip() for name in names_raw.replace("\r", "").split("\n")]
    names = [name for name in names if name]
    if not names:
        flash("请输入至少一个有效的班级名称", "danger")
        return redirect(url_for("readers.create_class"))

    created_count = 0
    skipped: List[str] = []
    for name in names:
        exists = Class.query.filter_by(name=name, grade_id=grade.id, is_deleted=False).first()
        if exists:
            skipped.append(name)
            continue
        klass = Class(name=name, grade_id=grade.id)
        db.session.add(klass)
        created_count += 1

    if created_count == 0:
        db.session.rollback()
        flash("没有新增班级，可能全部已存在。", "warning")
        return redirect(url_for("readers.create_class"))

    db.session.commit()
    message = f"成功创建 {created_count} 个班级"
    if skipped:
        message += f"，以下名称已存在并被跳过：{', '.join(skipped)}"
    flash(message, "success" if created_count else "warning")
    return redirect(url_for("readers.manage_classes"))


@bp.route("/classes/<int:class_id>/delete", methods=["POST"])
@login_required
def delete_class(class_id: int):
    if current_user.level != "admin":
        flash("只有管理员可以执行删除操作", "danger")
        return redirect(url_for("readers.manage_classes"))
    klass = Class.query.get_or_404(class_id)
    klass.is_deleted = True
    db.session.commit()
    flash("班级已删除", "success")
    return redirect(url_for("readers.manage_classes"))
