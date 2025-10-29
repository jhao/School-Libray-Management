import io
from datetime import datetime
from importlib import import_module
from importlib.util import find_spec
from typing import TYPE_CHECKING, List, Optional, Tuple

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from markupsafe import Markup
from sqlalchemy.orm import joinedload

if TYPE_CHECKING:  # pragma: no cover - used solely for type checkers
    from openpyxl import Workbook as WorkbookType
    from openpyxl import load_workbook as LoadWorkbookType
else:  # pragma: no cover - executed at runtime but simply provides aliases
    WorkbookType = object
    LoadWorkbookType = object

_OPENPYXL_CACHE: Tuple[Optional[WorkbookType], Optional[LoadWorkbookType], bool]
_OPENPYXL_CACHE = (None, None, False)


def _ensure_openpyxl() -> Tuple[Optional[WorkbookType], Optional[LoadWorkbookType], bool]:
    """Attempt to import openpyxl lazily.

    Returns the Workbook class, load_workbook callable and a boolean indicating
    whether the dependency is available. The result is cached to avoid repeated
    import attempts during a single request lifecycle.
    """

    global _OPENPYXL_CACHE
    workbook_cls, load_workbook_fn, has_attempted = _OPENPYXL_CACHE
    if has_attempted:
        return workbook_cls, load_workbook_fn, workbook_cls is not None and load_workbook_fn is not None

    if find_spec("openpyxl") is None:  # pragma: no cover - optional dependency may be absent
        _OPENPYXL_CACHE = (None, None, True)
        return None, None, False

    module = import_module("openpyxl")

    workbook_cls = getattr(module, "Workbook", None)
    load_workbook_fn = getattr(module, "load_workbook", None)
    _OPENPYXL_CACHE = (workbook_cls, load_workbook_fn, True)
    return workbook_cls, load_workbook_fn, workbook_cls is not None and load_workbook_fn is not None

from ..extensions import db
from ..models import Book, Category
from ..utils.category_tree import build_category_tree, flatten_category_tree
from ..utils.pagination import get_page_args


bp = Blueprint("books", __name__, url_prefix="/books")


def _category_options():
    categories = (
        Category.query.filter_by(is_deleted=False)
        .order_by(Category.sort, Category.name)
        .all()
    )
    category_tree = build_category_tree(categories)
    return list(flatten_category_tree(category_tree))


@bp.route("/")
@login_required
def list_books():
    keyword = request.args.get("q", "").strip()
    call_number = request.args.get("call_number", "").strip()
    position = request.args.get("position", "").strip()
    category_id_raw = request.args.get("category_id", "").strip()
    category_id = int(category_id_raw) if category_id_raw.isdigit() else None
    page, per_page = get_page_args()
    query = (
        Book.query.filter_by(is_deleted=False)
        .options(joinedload(Book.category), joinedload(Book.updated_by))
    )
    if keyword:
        like = f"%{keyword}%"
        query = query.filter((Book.name.like(like)) | (Book.isbn.like(like)))
    if call_number:
        query = query.filter(Book.call_number.like(f"%{call_number}%"))
    if position:
        query = query.filter(Book.position.like(f"%{position}%"))
    if category_id is not None:
        query = query.filter(Book.category_id == category_id)
    pagination = query.order_by(Book.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template(
        "books/list.html",
        books=pagination.items,
        filters={
            "keyword": keyword,
            "call_number": call_number,
            "position": position,
            "category_id": category_id,
        },
        pagination=pagination,
        category_options=_category_options(),
    )


@bp.route("/create", methods=["GET", "POST"])
@login_required
def create_book():
    if request.method == "GET":
        category_options = _category_options()
        return render_template("books/create.html", category_options=category_options)

    form = request.form
    name = form.get("name", "").strip()
    if not name:
        flash("图书名称不能为空", "danger")
        return redirect(url_for("books.create_book"))

    isbn = form.get("isbn", "").strip()
    if not isbn:
        flash("ISBN不能为空", "danger")
        return redirect(url_for("books.create_book"))

    amount_raw = (form.get("amount") or "").strip()
    if not amount_raw:
        flash("数量不能为空", "danger")
        return redirect(url_for("books.create_book"))
    try:
        amount = int(amount_raw)
    except ValueError:
        flash("数量必须为数字", "danger")
        return redirect(url_for("books.create_book"))
    if amount <= 0:
        flash("数量必须大于0", "danger")
        return redirect(url_for("books.create_book"))

    call_number = form.get("call_number") or None
    position = form.get("position")
    category_id_value = int(form.get("category_id")) if form.get("category_id") else None
    lend_amount = int(form.get("lend_amount", 0) or 0)
    price = form.get("price") or 0
    publisher = form.get("publisher")
    author = form.get("author")
    version = form.get("version")
    source = form.get("source")
    index_id = form.get("index_id")
    pages = form.get("pages") or None
    images = form.get("images")
    summary = form.get("summary")
    input_num = form.get("input_num") or None
    remark = form.get("remark")

    existing_active = Book.query.filter_by(isbn=isbn, is_deleted=False).first()
    if existing_active:
        flash("该ISBN已存在", "danger")
        return redirect(url_for("books.create_book"))

    existing_deleted = Book.query.filter_by(isbn=isbn, is_deleted=True).first()
    if existing_deleted:
        existing_deleted.name = name
        existing_deleted.isbn = isbn
        existing_deleted.call_number = call_number
        existing_deleted.position = position
        existing_deleted.category_id = category_id_value
        existing_deleted.amount = amount
        existing_deleted.lend_amount = lend_amount
        existing_deleted.price = price
        existing_deleted.publisher = publisher
        existing_deleted.author = author
        existing_deleted.version = version
        existing_deleted.source = source
        existing_deleted.index_id = index_id
        existing_deleted.pages = pages
        existing_deleted.images = images
        existing_deleted.summary = summary
        existing_deleted.input_num = input_num
        existing_deleted.remark = remark
        existing_deleted.updated_by_id = current_user.id
        existing_deleted.is_deleted = False
        db.session.commit()
        flash("图书创建成功", "success")
        return redirect(url_for("books.list_books"))

    book = Book(
        name=name,
        isbn=isbn,
        call_number=call_number,
        position=position,
        category_id=category_id_value,
        amount=amount,
        lend_amount=lend_amount,
        price=price,
        publisher=publisher,
        author=author,
        version=version,
        source=source,
        index_id=index_id,
        pages=pages,
        images=images,
        summary=summary,
        input_num=input_num,
        remark=remark,
    )
    book.updated_by_id = current_user.id
    db.session.add(book)
    db.session.commit()
    flash("图书创建成功", "success")
    return redirect(url_for("books.list_books"))


@bp.route("/<int:book_id>/update", methods=["POST"])
@login_required
def update_book(book_id: int):
    book = Book.query.get_or_404(book_id)
    form = request.form
    name = form.get("name", "").strip()
    isbn = form.get("isbn", "").strip()
    amount_raw = (form.get("amount") or "").strip()

    if not name:
        flash("图书名称不能为空", "danger")
        return redirect(url_for("books.list_books"))
    if not isbn:
        flash("ISBN不能为空", "danger")
        return redirect(url_for("books.list_books"))
    if not amount_raw:
        flash("数量不能为空", "danger")
        return redirect(url_for("books.list_books"))

    try:
        amount = int(amount_raw)
    except ValueError:
        flash("数量必须为数字", "danger")
        return redirect(url_for("books.list_books"))

    if amount <= 0:
        flash("数量必须大于0", "danger")
        return redirect(url_for("books.list_books"))

    book.name = name
    book.isbn = isbn
    book.call_number = form.get("call_number") or None
    book.position = form.get("position")
    book.category_id = int(form.get("category_id")) if form.get("category_id") else None
    book.amount = amount
    book.lend_amount = int(form.get("lend_amount", book.lend_amount) or book.lend_amount)
    book.price = form.get("price") or book.price
    book.publisher = form.get("publisher")
    book.author = form.get("author")
    book.version = form.get("version")
    book.source = form.get("source")
    book.index_id = form.get("index_id")
    book.pages = form.get("pages") or book.pages
    book.images = form.get("images")
    book.summary = form.get("summary")
    book.input_num = form.get("input_num") or book.input_num
    book.remark = form.get("remark")
    book.updated_by_id = current_user.id
    db.session.commit()
    flash("图书信息已更新", "success")
    return redirect(url_for("books.list_books"))


@bp.route("/<int:book_id>/delete", methods=["POST"])
@login_required
def delete_book(book_id: int):
    if current_user.level != "admin":
        flash("只有管理员可以执行删除操作", "danger")
        return redirect(url_for("books.list_books"))
    book = Book.query.get_or_404(book_id)
    book.is_deleted = True
    book.updated_by_id = current_user.id
    db.session.commit()
    flash("图书已删除", "success")
    return redirect(url_for("books.list_books"))


@bp.route("/import", methods=["POST"])
@login_required
def import_books():
    _, load_workbook_fn, has_openpyxl = _ensure_openpyxl()
    if not has_openpyxl or load_workbook_fn is None:
        flash("未安装 openpyxl 库，无法导入。请先运行 pip install openpyxl。", "danger")
        return redirect(url_for("books.list_books"))

    if request.form.get("template_confirmed") != "1":
        flash("请先下载导入模板并确认后再上传数据。", "warning")
        return redirect(url_for("books.list_books"))

    file = request.files.get("file")
    if not file:
        flash("请选择Excel文件", "danger")
        return redirect(url_for("books.list_books"))

    try:
        wb = load_workbook_fn(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created_count = 0
        restored_count = 0
        skipped: List[str] = []

        categories = (
            Category.query.filter_by(is_deleted=False)
            .order_by(Category.sort, Category.name)
            .all()
        )
        category_tree = build_category_tree(categories)
        flattened_categories = list(flatten_category_tree(category_tree))
        category_display_map = {}
        category_by_name = {}
        category_by_id = {str(category.id): category for category in categories}
        for item in flattened_categories:
            category = item["category"]
            depth = item["depth"]
            prefix = "└ " if depth > 0 else ""
            indent = "  " * depth
            display = f"{indent}{prefix}{category.name}"
            keys = {display, display.strip(), category.name}
            for key in keys:
                if key:
                    category_display_map.setdefault(key, category)
            category_by_name.setdefault(category.name, []).append(category)

        for index, row in enumerate(rows, start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            name_value = str(row[0]).strip() if len(row) > 0 and row[0] else "未命名图书"
            isbn = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not isbn:
                skipped.append(f"第{index}行: ISBN 不能为空")
                continue

            existing_active = Book.query.filter_by(isbn=isbn, is_deleted=False).first()
            if existing_active:
                skipped.append(f"第{index}行(ISBN {isbn}): 已存在未删除的图书")
                continue

            category_raw = str(row[2]) if len(row) > 2 and row[2] is not None else ""
            category_value = category_raw.strip() if category_raw else ""
            call_number_raw = row[3] if len(row) > 3 else None
            call_number = (
                str(call_number_raw).strip() if call_number_raw not in (None, "") else None
            )
            position_raw = row[4] if len(row) > 4 else None
            position = (
                str(position_raw).strip() if position_raw not in (None, "") else None
            )
            raw_amount = row[5] if len(row) > 5 else None
            try:
                if raw_amount in (None, ""):
                    amount = 1
                else:
                    amount = int(float(raw_amount))
            except (TypeError, ValueError):
                skipped.append(f"第{index}行(ISBN {isbn}): 数量格式不正确")
                continue
            if amount <= 0:
                skipped.append(f"第{index}行(ISBN {isbn}): 数量必须大于0")
                continue

            price = row[6] if len(row) > 6 else 0
            publisher = row[7] if len(row) > 7 else None
            author = row[8] if len(row) > 8 else None
            summary = row[9] if len(row) > 9 else None

            category_obj = None
            if category_raw:
                category_obj = category_display_map.get(category_raw) or category_display_map.get(
                    category_value
                )
            if category_obj is None and category_value:
                category_obj = category_by_id.get(category_value)
                if category_obj is None:
                    try:
                        category_numeric = int(float(category_value))
                    except (TypeError, ValueError):
                        category_numeric = None
                    if category_numeric is not None:
                        category_obj = category_by_id.get(str(category_numeric))
                if category_obj is None:
                    matches = category_by_name.get(category_value)
                    if matches:
                        category_obj = matches[0]

            existing_deleted = Book.query.filter_by(isbn=isbn, is_deleted=True).first()
            if existing_deleted:
                existing_deleted.name = name_value or existing_deleted.name or "未命名图书"
                existing_deleted.call_number = call_number
                existing_deleted.position = position
                existing_deleted.amount = amount
                existing_deleted.lend_amount = 0
                existing_deleted.price = price or 0
                existing_deleted.publisher = publisher
                existing_deleted.author = author
                existing_deleted.summary = summary
                existing_deleted.updated_by_id = current_user.id
                existing_deleted.category_id = category_obj.id if category_obj else None
                existing_deleted.is_deleted = False
                restored_count += 1
                continue

            book = Book(
                name=name_value,
                isbn=isbn,
                call_number=call_number,
                position=position,
                amount=amount,
                price=price or 0,
                publisher=publisher,
                author=author,
                summary=summary,
                category_id=category_obj.id if category_obj else None,
            )
            book.updated_by_id = current_user.id
            db.session.add(book)
            created_count += 1

        db.session.commit()

        success_parts = []
        if created_count:
            success_parts.append(f"成功导入 {created_count} 条图书记录")
        if restored_count:
            success_parts.append(f"恢复 {restored_count} 条已删除图书记录")
        if success_parts:
            flash("，".join(success_parts), "success")

        if skipped:
            preview = skipped[:100]
            if len(skipped) > 100:
                preview.append(f"……共 {len(skipped)} 条错误，仅显示前100条")
            flash(Markup("部分数据导入失败：<br>" + "<br>".join(preview)), "danger")
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        flash(f"导入失败: {exc}", "danger")

    return redirect(url_for("books.list_books"))


@bp.route("/import-template")
@login_required
def download_import_template():
    workbook_cls, _, has_openpyxl = _ensure_openpyxl()
    if not has_openpyxl or workbook_cls is None:
        flash("未安装 openpyxl 库，无法生成模板。请先运行 pip install openpyxl。", "danger")
        return redirect(url_for("books.list_books"))

    wb = workbook_cls()
    ws = wb.active
    ws.title = "图书信息"
    ws.append([
        "图书名称",
        "ISBN",
        "分类",
        "索书码",
        "位置",
        "数量",
        "价格",
        "出版社",
        "作者",
        "简介",
    ])

    categories = (
        Category.query.filter_by(is_deleted=False)
        .order_by(Category.sort, Category.name)
        .all()
    )
    category_tree = build_category_tree(categories)
    flattened_categories = list(flatten_category_tree(category_tree))

    data_sheet = wb.create_sheet("分类信息")
    data_sheet.append(["分类", "分类ID", "层级", "原始名称"])
    for idx, item in enumerate(flattened_categories, start=2):
        category = item["category"]
        depth = item["depth"]
        prefix = "└ " if depth > 0 else ""
        indent = "  " * depth
        display_name = f"{indent}{prefix}{category.name}"
        data_sheet.cell(row=idx, column=1, value=display_name)
        data_sheet.cell(row=idx, column=2, value=category.id)
        data_sheet.cell(row=idx, column=3, value=depth)
        data_sheet.cell(row=idx, column=4, value=category.name)
    data_sheet.sheet_state = "hidden"

    try:
        dv_module = import_module("openpyxl.worksheet.datavalidation")
        DataValidation = getattr(dv_module, "DataValidation", None)
        utils_module = import_module("openpyxl.utils")
        get_column_letter = getattr(utils_module, "get_column_letter", None)
        quote_sheetname_fn = getattr(utils_module, "quote_sheetname", None)
    except Exception:  # noqa: BLE001 - fallback when openpyxl structure changes
        DataValidation = None
        get_column_letter = None
        quote_sheetname_fn = None

    def _quote_sheet_name(name: str) -> str:
        if quote_sheetname_fn:
            return quote_sheetname_fn(name)
        escaped = name.replace("'", "''")
        return f"'{escaped}'"

    if DataValidation and get_column_letter:
        max_rows = 500
        category_column = get_column_letter(3)
        if flattened_categories:
            category_list_ref = (
                f"={_quote_sheet_name(data_sheet.title)}!$A$2:$A${len(flattened_categories) + 1}"
            )
            category_range = f"{category_column}2:{category_column}{max_rows}"
            category_validation = DataValidation(
                type="list",
                formula1=category_list_ref,
                allow_blank=True,
            )
            category_validation.error = "请选择下拉列表中的分类"
            category_validation.errorTitle = "无效的分类"
            ws.add_data_validation(category_validation)
            category_validation.add(category_range)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="book-import-template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/export")
@login_required
def export_books():
    workbook_cls, _, has_openpyxl = _ensure_openpyxl()
    if not has_openpyxl or workbook_cls is None:
        flash("未安装 openpyxl 库，无法导出。请先运行 pip install openpyxl。", "danger")
        return redirect(url_for("books.list_books"))

    wb = workbook_cls()
    ws = wb.active
    ws.append([
        "图书名称",
        "ISBN",
        "位置",
        "总数量",
        "已借出",
        "价格",
        "出版社",
        "作者",
        "简介",
    ])
    for book in Book.query.filter_by(is_deleted=False).order_by(Book.name).all():
        ws.append(
            [
                book.name,
                book.isbn,
                book.position,
                book.amount,
                book.lend_amount,
                float(book.price or 0),
                book.publisher,
                book.author,
                book.summary,
            ]
        )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"books-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
